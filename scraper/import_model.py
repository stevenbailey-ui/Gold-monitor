#!/usr/bin/env python3
"""
Local model importer — v22 layout.

Reads the portfolio model (.xlsx) and writes data/model_snapshot.json: the valuation
layer the scraper merges into the public dashboard (NAV £/share, the scenario £m grid,
the 2029 gold deck, and 2029 exit income/yield).

Runs LOCALLY only — never in the GitHub Action — so the spreadsheet never leaves your
machine. Requires openpyxl.

    python scraper/import_model.py /path/to/portfolio_v22.xlsx

CHANGED FROM THE v21 IMPORTER
-----------------------------
v21 was read from Summary / NPV Model / Assumptions. v22 has none of those sheets, so
every cell reference below is new. Specifically:

  * Scenario column order INVERTED. v21 Summary ran B=bull C=base D=bear.
    v22 runs B=Bear C=Base D=Bull. Reading the old order silently swaps bull and bear.
  * Discount rates now live on Inputs, and are EQUAL at 6% for both holdings. The old
    8%/12% split double-counted jurisdiction risk already carried by the delivery
    probabilities, the seizure residual and the P/NAV path.
  * v22 is NAV-only — there is no P/FCF leg. The `pfcf_mult` key is retained for
    front-end compatibility and now carries the 2030 terminal P/NAV multiple, which is
    where jurisdiction risk is expressed under the new framing. npv_wt is pinned to 1.0
    and pfcf_wt to 0.0.
  * 2029 income is no longer a model cell. It is computed here from the 2029 base
    portfolio value and the exit_plan block in data/manual_inputs.json.

Cells that read #VALUE!/blank are skipped and the previous snapshot value is kept.
Open + recalculate the model before importing.
"""
import json, os, sys, datetime

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
DATA = os.path.join(ROOT, "data")
SNAP = os.path.join(DATA, "model_snapshot.json")
MANUAL = os.path.join(DATA, "manual_inputs.json")

# Anonymised holding keys. No tickers, no jurisdictions, no share counts in this file.
HOLD = ["africa", "asia"]

# --- v22 cell map -----------------------------------------------------------
# Scenario NAV per share (£). Columns B/C/D = Bear/Base/Bull.
NAV_SHEET = {"africa": "THX_NAV", "asia": "MTL_NAV"}
NAV_ROW = 67
NAV_COL = {"bear": "B", "base": "C", "bull": "D"}

# Inputs sheet scalars.
DISC_CELL = {"africa": "B15", "asia": "B16"}
TERM_PNAV = {"africa": "B65", "asia": "C65"}      # 2030 P/NAV — risk now lives here
ANCHOR_PNAV = {"africa": "B61", "asia": "C61"}    # 2026 P/NAV, auto-calibrated to price

# Portfolio sheet: scenario value grid, £. Rows Bear/Base/Bull, cols B..F = 2026..2030.
SCEN_ROW = {"bear": 20, "base": 21, "bull": 22}
SCEN_COL = {"2026": "B", "2027": "C", "2028": "D", "2029": "E", "2030": "F"}

# Forecast sheet: gold deck by scenario. Rows 5/6/7 = Bear/Base/Bull, col E = 2029.
GOLD_ROW = {"bear": 5, "base": 6, "bull": 7}
GOLD_2029_COL = "E"

# Valuation sheet: rotation trigger.
CROSS_CELL, CROSS_THRESH_SHEET, CROSS_THRESH_CELL = "B15", "Inputs", "B75"

FELL_BACK = []


def num(ws, coord, prior, allow_zero=True):
    v = ws[coord].value
    if isinstance(v, (int, float)) and (allow_zero or v != 0):
        return float(v)
    FELL_BACK.append(f"{ws.title}!{coord}")
    return prior


def main():
    if len(sys.argv) < 2:
        sys.exit("usage: python scraper/import_model.py /path/to/portfolio_v22.xlsx")
    try:
        import openpyxl
    except ImportError:
        sys.exit("openpyxl required:  pip install openpyxl")

    path = sys.argv[1]
    wb = openpyxl.load_workbook(path, data_only=True)

    missing = [s for s in ("Inputs", "THX_NAV", "MTL_NAV", "Forecast", "Portfolio", "Valuation")
               if s not in wb.sheetnames]
    if missing:
        sys.exit(f"not a v22 model — missing sheet(s): {', '.join(missing)}. "
                 "The v21 importer read Summary/NPV Model/Assumptions; that map no longer applies.")

    inp = wb["Inputs"]; fc = wb["Forecast"]; pf = wb["Portfolio"]; vl = wb["Valuation"]

    prior = json.load(open(SNAP)) if os.path.exists(SNAP) else {}
    pv = prior.get("valuation", {})
    ps = prior.get("scenarios", {})

    val = {}
    for key in HOLD:
        ws = wb[NAV_SHEET[key]]
        p = pv.get(key, {})
        val[key] = {
            "npv_bull": round(num(ws, f"{NAV_COL['bull']}{NAV_ROW}", p.get("npv_bull")), 3),
            "npv_base": round(num(ws, f"{NAV_COL['base']}{NAV_ROW}", p.get("npv_base")), 3),
            "npv_bear": round(num(ws, f"{NAV_COL['bear']}{NAV_ROW}", p.get("npv_bear")), 3),
            "disc": round(num(inp, DISC_CELL[key], p.get("disc")), 4),
            # Retained key, new meaning: terminal P/NAV, not a P/FCF multiple.
            "pfcf_mult": round(num(inp, TERM_PNAV[key], p.get("pfcf_mult")), 3),
            "pnav_anchor": round(num(inp, ANCHOR_PNAV[key], p.get("pnav_anchor")), 3),
            "npv_wt": 1.0,
            "pfcf_wt": 0.0,
        }

    scen = {}
    for yr, col in SCEN_COL.items():
        pp = ps.get(yr, {})
        row = {}
        for s in ("bear", "base", "bull"):
            v = num(pf, f"{col}{SCEN_ROW[s]}", None)
            row[s] = round(v / 1e6, 2) if v is not None else pp.get(s)
        scen[yr] = row

    gold_fc = num(fc, f"{GOLD_2029_COL}{GOLD_ROW['base']}", prior.get("gold_2029_forecast"),
                  allow_zero=False)

    # 2029 exit income, computed here rather than read from a model cell.
    inc = prior.get("income_2029_base")
    yld = prior.get("income_2029_yield")
    try:
        man = json.load(open(MANUAL))
        ep = man.get("exit_plan", {})
        share = float(ep.get("transfer_share", 0.5))
        fyield = float(ep.get("fund_yield", 0.045))
        miner_yield = float(ep.get("miner_net_yield", 0.037))
        v29 = scen["2029"]["base"]
        if v29:
            gross = v29 * 1e6 * (share * fyield + (1 - share) * miner_yield)
            inc = int(round(gross))
            yld = round(100 * gross / (v29 * 1e6), 2)
    except Exception as e:
        FELL_BACK.append(f"exit_plan ({e})")

    cross = num(vl, CROSS_CELL, None)
    thresh = num(inp, CROSS_THRESH_CELL, None)

    out = {
        "model_as_of": datetime.date.today().isoformat(),
        "source": os.path.basename(path),
        "schema": "v22",
        "valuation": val,
        "scenarios": scen,
        "gold_2029_forecast": round(gold_fc) if gold_fc else prior.get("gold_2029_forecast"),
        "income_2029_base": inc,
        "income_2029_yield": yld,
        "rotation": {
            "cross_ratio": round(cross, 3) if cross else None,
            "threshold": round(thresh, 3) if thresh else None,
            "signal": ("ROTATE" if (cross and thresh and cross >= thresh) else "HOLD")
                      if (cross and thresh) else None,
        },
        "method_note": (
            "Both holdings discounted at a common 6%. Jurisdiction and delivery risk are "
            "expressed once, in the delivery probability and the P/NAV path, not in the rate."
        ),
    }
    json.dump(out, open(SNAP, "w"), indent=2)

    n = len(FELL_BACK)
    print(f"wrote model_snapshot.json from {out['source']} (schema v22) — "
          + ("all cells resolved." if n == 0 else
             f"{n} unresolved: {', '.join(FELL_BACK[:6])}. Recalculate the model and re-run."))
    if out["rotation"]["signal"]:
        print(f"  rotation: cross {out['rotation']['cross_ratio']} vs threshold "
              f"{out['rotation']['threshold']} -> {out['rotation']['signal']}")


if __name__ == "__main__":
    main()
